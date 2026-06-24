"""
qualitative.py — reproduce the coded open-ended evidence (Figures 8-9, Table 3).

The hand-coded ground truth is the team's coding pack (the same artifact the
thesis used). It is parsed into a compact, quote-free table of
(session_prefix, version, item, codes) so the figures regenerate without
shipping participant quotes. Counts are then restricted to the 32-participant
analysis cohort and split by arm (Shared/main+direct = integrated; Baseline).

Codes (codebook): P1 resonance, P2 vivid scene, P3 continuity bridge,
P4 voice match, P5 agency; F1-F7 friction (F3 = immersion break dominates);
N1-N4 unmet need (depth / breadth / guidance / practical next step).
"""
from __future__ import annotations
import json
import re
from collections import Counter
from pathlib import Path

_PKG = Path(__file__).resolve().parent
_CODES = _PKG.parent / "data" / "coding_pack_codes.json"

POSITIVE = ["P1", "P2", "P3", "P4", "P5"]
POSITIVE_LABEL = {"P1": "Resonance", "P2": "Vivid scene", "P3": "Continuity",
                  "P4": "Voice match", "P5": "Agency"}
_CODE_RE = re.compile(r"\b([PFN][1-7])\b", re.I)


def parse_codes(cell):
    """All distinct codes mentioned in a coder cell ('P1 & P5 / P3' -> P1,P3,P5)."""
    if cell is None:
        return []
    return sorted({m.upper() for m in _CODE_RE.findall(str(cell))})


# --------------------------------------------------------------------------- #
# load the coded table                                                         #
# --------------------------------------------------------------------------- #
def load_coding_table(xlsx_path=None):
    """Return (coding_rows, reliability_pairs) from the pack.

    coding_rows: list of {sheet, session, version, item, codes[]}
    reliability_pairs: list of {coder_a, coder_b}

    Prefers the shipped quote-free JSON; falls back to parsing the xlsx if a
    path is given.
    """
    if _CODES.exists() and not xlsx_path:
        d = json.loads(_CODES.read_text())
        return d["open_ended_and_resonance"], d["reliability_pairs"]
    return _parse_xlsx(xlsx_path)


def _parse_xlsx(xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    out = []
    # Open-ended feedback: rows with an int in col 0 are data rows
    ws = wb["Open-ended feedback"]
    for r in ws.iter_rows(values_only=True):
        if r and isinstance(r[0], int):
            item = "felt_real" if "real" in str(r[3] or "").lower() else "broke"
            out.append({"sheet": "open_ended", "session": str(r[1])[:8],
                        "version": str(r[2] or "").strip(), "item": item,
                        "codes": parse_codes(r[5])})
    # Resonance candidates: substantive replies coded for positive signal
    ws = wb["Resonance candidates"]
    for r in ws.iter_rows(values_only=True):
        if r and isinstance(r[0], int):
            out.append({"sheet": "resonance", "session": str(r[1])[:8],
                        "version": str(r[2] or "").strip(), "item": "transcript",
                        "codes": parse_codes(r[5])})
    # Reliability sample: two coders -> for kappa
    ws = wb["Reliability sample (kappa)"]
    rel = []
    for r in ws.iter_rows(values_only=True):
        if r and isinstance(r[0], int):
            rel.append({"coder_b": _norm(r[5]), "coder_a": _norm(r[6])})
    return out, rel


def _norm(v):
    s = str(v or "").strip().upper()
    return "NONE" if s in ("N", "NONE", "") else s


# --------------------------------------------------------------------------- #
# cohort map                                                                   #
# --------------------------------------------------------------------------- #
def cohort_arm_map(sample):
    """8-char session prefix -> arm, for the 32-participant analysis cohort."""
    return {s["meta"]["sessionId"][:8]: s["_arm"] for s in sample}


# --------------------------------------------------------------------------- #
# Figure 8 (left): positive codes on 'felt real' by arm                        #
# --------------------------------------------------------------------------- #
def felt_real_codes_by_arm(coding_table, sample):
    armmap = cohort_arm_map(sample)
    counts = {"integrated": Counter(), "baseline": Counter()}
    for row in coding_table:
        if row["sheet"] != "open_ended" or row["item"] != "felt_real":
            continue
        arm = armmap.get(row["session"])
        if arm is None:
            continue
        for c in row["codes"]:
            if c in POSITIVE:
                counts[arm][c] += 1
    return counts


# --------------------------------------------------------------------------- #
# Figure 8 (right): open-ended word counts by arm (from the DB text)           #
# --------------------------------------------------------------------------- #
def open_ended_word_counts(sample):
    def wc(t):
        return len((t or "").split())
    res = {}
    for arm in ["integrated", "baseline"]:
        sub = [s for s in sample if s["_arm"] == arm]
        n = len(sub)
        # Each session contributes (wrote_real_flag, words_real, words_broke).
        # The bundled snapshot ships these as precomputed numeric metrics
        # (postSurvey._oe) instead of raw free text; otherwise compute from text.
        wrote = words_r = words_b = 0
        for s in sub:
            oe = (s.get("postSurvey") or {}).get("_oe")
            if isinstance(oe, dict):
                wrote += int(oe.get("wrote_real", 0))
                words_r += oe.get("words_real", 0)
                words_b += oe.get("words_broke", 0)
            else:
                r = (s.get("postSurvey") or {}).get("oe_real", "") or ""
                b = (s.get("postSurvey") or {}).get("oe_broke", "") or ""
                wrote += int(bool(r.strip()))
                words_r += wc(r)
                words_b += wc(b)
        # "Mean words written per open-ended item" (Fig 8 right) divides total
        # words by the arm size, so a non-writer contributes 0 words.
        res[arm] = {
            "n": n,
            "wrote_felt_real": wrote,
            "pct_felt_real": 100 * wrote / n if n else 0,
            "mean_words_real": words_r / n if n else 0,
            "mean_words_broke": words_b / n if n else 0,
        }
    return res


# --------------------------------------------------------------------------- #
# Figure 9: requests for concrete career specifics                             #
# --------------------------------------------------------------------------- #
# The thesis reports the human-coded totals (Skills 18 / Direction 14 / Pay 4).
# Those coded counts are the canonical Figure-9 values (HUMAN_CODED below) and
# are what the figure renders. requests_for_specifics() is a transparent
# automated PROXY over the transcripts that corroborates the ordering (skills &
# direction common; explicit pay questions rarest); it is not a substitute for
# the hand coding, which can tell a genuine "what's the salary?" request from an
# incidental mention of money.
HUMAN_CODED_REQUESTS = {"skills": 18, "direction": 14, "pay": 4}

# Explicit, question-shaped money asks only (keeps the proxy's pay count honest).
_PAY = re.compile(r"(salary|how much (can|do|will|would) (i|you)|"
                  r"\b(money|income|earn\w*|pay|paid|wage)\b[^.?!]{0,40}\?|"
                  r"\bhonest(ly)?\b[^.?!]{0,15}\bmoney\b)", re.I)
_SKILL = re.compile(r"\b(skill|skills|learn|study|studying|course|courses|"
                    r"build|practice|master|improve|prepare)\b", re.I)
_DIRECTION = re.compile(r"\b(where (do|should|to) (i|you)|how (do|did|should) (i|you)|"
                        r"first job|next step|which (career|role|field|path)|"
                        r"right (career|choice|path|for me))\b", re.I)


def requests_for_specifics(sample):
    """Automated proxy (per-participant presence) — corroborates Figure 9's
    ordering only; the canonical values are HUMAN_CODED_REQUESTS."""
    cat = {"skills": 0, "direction": 0, "pay": 0}
    for s in sample:
        asked = {"skills": False, "direction": False, "pay": False}
        for m in (s.get("phaseC") or {}).get("transcript", []):
            if (m or {}).get("role") != "user":
                continue
            t = m.get("text", "")
            if _PAY.search(t):
                asked["pay"] = True
            elif _SKILL.search(t):
                asked["skills"] = True
            elif _DIRECTION.search(t):
                asked["direction"] = True
        for k, v in asked.items():
            cat[k] += int(v)
    return cat


# --------------------------------------------------------------------------- #
# inter-rater reliability (Cohen's kappa)                                       #
# --------------------------------------------------------------------------- #
def cohen_kappa(pairs):
    a = [p["coder_a"] for p in pairs]
    b = [p["coder_b"] for p in pairs]
    n = len(a)
    po = sum(1 for x, y in zip(a, b) if x == y) / n
    cats = set(a) | set(b)
    pe = sum((a.count(c) / n) * (b.count(c) / n) for c in cats)
    kappa = (po - pe) / (1 - pe) if pe < 1 else 1.0
    return {"n": n, "observed_agreement": po, "expected_agreement": pe, "kappa": kappa}
